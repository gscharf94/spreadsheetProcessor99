from openpyxl import Workbook, load_workbook
from itertools import permutations

class Handler():
	def __init__(self,fileName):
		### takes in import file
		self.fileName = fileName
		self.wb = load_workbook(fileName)
		self.sheetNames = self.wb.sheetnames

	def getRecordsFromSheet(self,sheetName):
		sheet = self.wb[sheetName]
		edge = self.findEdge(sheet)

		rows = []
		
		cols = ['A','B','C','D','E','F','G','H']

		for x in range(2,edge):
			row = []
			for y in cols:
				loc = y+str(x)
				row.append(sheet[loc].value)
			rows.append(row)

		return rows

	def findEdge(self,sheet):
		counter = 1
		while True:
			if sheet['A'+str(counter)].value == None:
				break
			else:
				counter += 1
		return counter

	def initMainDict(self):
		canvNames = self.sheetNames[2:]
		self.mainDict = {}
		for name in canvNames:
			rows = self.getRecordsFromSheet(name)
			self.mainDict[name] = rows

	def initExtraDict(self):
		self.extraDict = {}
		canvNames = self.sheetNames[2:]
		for name in canvNames:
			self.extraDict[name] = []

		sheet = self.wb[self.sheetNames[0]]
		edge = self.findEdge(sheet)

		cols = ['B','D','F']
		rows = []
		for x in range(1,edge):
			row = []
			for y in cols:
				loc = y+str(x)
				row.append(sheet[loc].value)
			rows.append(row)
		for row in rows:
			### special case because this name is too long to fit into an excel sheetname
			if row[2] == 'Sisepuede2020, Fvawcanvasser21@gmai':
				self.extraDict['Sisepuede2020'].append(row[0:2])
			else:
				self.extraDict[row[2]].append(row[0:2])
	def convertName(self,name):
		### takes a name and returns it better ;)
		split = name.split(",")
		# print(split)
		split2 = split[1].split(" ")
		# print(split2)
		newName = f'{split[0]}, {split2[1]}'
		# print(newName)
		return newName

	def getPotentialCombinations(self,name):
		initialName = name
		potentialNames = [name]
		name = name.replace(".","")
		potentialNames.append(name)

		split = name.split(",")
		lastName = split[0]

		firstNameParts = split[1].split(" ")[1:]
		possibleFirstNames = []


		if len(firstNameParts) == 3:
			perms = permutations(firstNameParts)
			for perm in perms:
				possibleFirstNames.append(perm)
			perms = permutations(firstNameParts,2)
			for perm in perms:
				possibleFirstNames.append(perm)
			for part in firstNameParts:
				possibleFirstNames.append(part)
		elif len(firstNameParts) == 4:
			perms = permutations(firstNameParts)
			for perm in perms:
				possibleFirstNames.append(perm)
			perms = permutations(firstNameParts,3)
			for perm in perms:
				possibleFirstNames.append(perm)
			perms = permutations(firstNameParts,2)
			for perm in perms:
				possibleFirstNames.append(perm)
			for part in firstNameParts:
				possibleFirstNames.append(part)
		elif len(firstNameParts) == 2:
			perms = permutations(firstNameParts)
			for perm in perms:
				possibleFirstNames.append(perm)
			for part in firstNameParts:
				possibleFirstNames.append(part)

		# print(f'lastName: {lastName} rest -> {firstNameParts}')
		names = []
		names.append(initialName)
		names.append(initialName.replace(".",""))
		for name in possibleFirstNames:
			outputStr = f'{lastName},'
			if type(name) == str:
				outputStr += f' {name}'
				names.append(outputStr)
			else:
				for chunk in name:
					outputStr += f' {chunk}'
					names.append(outputStr)
		return names

	def addDicts(self):
		c = 0
		for name in self.extraDict:
			couldNotFind = []
			rows = self.mainDict[name]
			responses = self.extraDict[name]
			for response in responses:
				name = self.convertName(response[0])
				for x,row in enumerate(rows):
					if row[1] == name and len(row)==8:
						if row[1] == 'Rodriguez, Maria':
							print(row)
						# print(f'found match {row[1]} - {name}')
						c+= 1
						row.append(response[1])
						break
					if x == len(rows)-1:
						# print(f'WE COULD NOT FIND {name} (original {response[0]}')
						couldNotFind.append(response)
			for chunk in couldNotFind:
				print(chunk)
				for x,row in enumerate(rows):
					potentialNames = self.getPotentialCombinations(chunk[0])
					for pot in potentialNames:
						if pot == row[1]:
							# print('found match for impossible')
							c += 1
							row.append(response[1])
							break
		# print(self.mainDict)
		print(c)

	def addDictss(self):
		name = 'Abreu, Rosaura'

		c = 0

		rows = self.mainDict[name]
		responses = self.extraDict[name]

		print(f'len(rows) {len(rows)} len(responses) {len(responses)}')
		for response in responses:
			name = self.convertName(response[0])
			for x,row in enumerate(rows):
				# print(f'comparing {row[1]} with {name}')
				if row[1] == name:
					c += 1
					print(f'--------\nfound match#{c} {row[1]} - {name}\n-----------')
					row.append(response[1])
					break
		print(self.mainDict['Abreu, Rosaura'])


handler = Handler('4-24-2020.xlsx')
# handler.getRecordsFromSheet('Abreu, Rosaura')
handler.initMainDict()
handler.initExtraDict()
handler.addDicts()

import pickle

pickle.dump(handler.mainDict,open('finalDict2.p','wb'))
