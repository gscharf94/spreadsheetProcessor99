import pickle
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side

GRAYFILL = PatternFill(fill_type="solid",start_color='00ededed',end_color='00ededed')
DARKGRAYFILL = PatternFill(fill_type="solid",start_color='00e0e0e0',end_color='00e0e0e0')


CENTERALIGN = Alignment(horizontal='center')
RIGHTALIGN = Alignment(horizontal='right')
LEFTALIGN = Alignment(horizontal='left')

ALLBORDER = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

RIGHTBORDER = Border(right=Side(style='thin'))
BOTTOMBORDER = Border(bottom=Side(style='thin'))
TOPBORDER = Border(top=Side(style='thin'))

TOPBOTTOMBORDER = Border(top=Side(style='thin'),bottom=Side(style='thin'))

BOLDFONT = Font(bold=True)

HEADERFONT = Font(size=12,bold=True)

TEAMS = {
		# 12pm-2pm y 3pm-5pm
		'Team Biaziris':
			[
			'Yomay, Sandra','Sanchez, Rixio','Notto, Guseppe',
			'Perdomo, Carlos','Pico, Maria Elizabeth','Molina, Biaziris'
			],
		# 2pm-4pm y 5pm-7pm
		'Team Eliana':
			[
			'Pulido, Sandra','Abreu, Rosaura','Ruiz, Sol',
			'Petit, Maria Alejandra','Conway, Miriam','Torres, Eliana'
			],
		# 2pm-4pm y 5pm-7pm
		'Team Zaki':
			[
			'Ruiz, Marisol','Chavez, Zoraima','Rosario, Arileida',
			'Rivera, Lizzy','Gallardo, Arianny','Banna, Zaki'
			],
		# 12 - 5
		'Group #1':
			[
			'Six, Fvaw','Eight, Fvaw','Ten, Fvaw','Fifteen, Fvaw',
			'Nine, Fvaw','Thirteen, Fvaw','Twelve, Fvaw','Fourteen, Fvaw'
			],
		# 1 - 6
		'Group #2':
			[
			'Twentytwo, Fvaw','Sisepuede2020, Fvawcanvasser21@gmai',
			'Five, Fvaw','Four, Fvaw','Nineteen, Fvaw','One, Fvaw',
			'Moreno Fuentes, Lazaraj','Two, Fvaw','Twenty, Fvaw'
			]
		}


class ExcelCreator():
	def __init__(self,importFile):
		self.mainDict = pickle.load(open(importFile,'rb'))
		### fixing special case
		if 'Sisepuede2020' in self.mainDict:
			val = self.mainDict.pop('Sisepuede2020')
			self.mainDict['Sisepuede2020, Fvawcanvasser21@gmai'] = val

	def createFile(self,group):
		canvs = TEAMS[group]
		wb = openpyxl.Workbook()

		### create first sheep, empty at first
		ws = wb.worksheets[0]
		ws.title = "Team Data"

		canvSheets = {}
		statsDict = {}

		### loop through each canvasser in group and set up their individual sheet
		### save an average list to add to team data page
		average = [0 for x in range(21)]

		workedToday = 0
		for canvName in canvs:
			if canvName in self.mainDict:
				workedToday += 1
				canvSheets[canvName] = wb.create_sheet(canvName)
				self.populateDetail(canvSheets[canvName],self.mainDict[canvName])
				stats = self.getIndividualStats(self.mainDict[canvName])
				statsDict[canvName] = stats
				average = self.addLists(average,stats)

		average = self.finishAverages(average,workedToday)

		self.writeTeamData(ws,statsDict,group,average)

		if "Group" in group:
			self.changeName(group,wb,canvSheets)

		self.saveFile(group,wb)

	def writeTeamData(self,ws,statsDict,group,average):
		ws.sheet_view.showGridLines = False
		statsDict['Averages'] = average

		ws.column_dimensions['A'].width = 22
		ws.column_dimensions['B'].width = 8
		ws.column_dimensions['C'].width = 11
		ws.column_dimensions['D'].width = 8
		ws.column_dimensions['E'].width = 10
		ws.column_dimensions['F'].width = 8
		ws.column_dimensions['G'].width = 10
		ws.column_dimensions['H'].width = 8
		ws.column_dimensions['I'].width = 10
		ws.column_dimensions['J'].width = 8
		ws.column_dimensions['K'].width = 12
		ws.column_dimensions['L'].width = 8
		ws.column_dimensions['M'].width = 12
		ws.column_dimensions['N'].width = 8
		ws.column_dimensions['O'].width = 16
		ws.column_dimensions['P'].width = 8
		ws.column_dimensions['Q'].width = 14
		ws.column_dimensions['R'].width = 8
		ws.column_dimensions['S'].width = 8
		ws.column_dimensions['T'].width = 8
		ws.column_dimensions['U'].width = 14
		ws.column_dimensions['V'].width = 8



		ws['A1'].value = "Canvasser"
		ws['A1'].font = HEADERFONT
		ws['A1'].alignment = RIGHTALIGN
		ws['B1'].value = "Total"
		ws['B1'].font = HEADERFONT
		ws['B1'].alignment = CENTERALIGN
		ws['B1'].border = BOTTOMBORDER
		ws['C1'].value = "Not Home"
		ws['C1'].font = HEADERFONT
		ws['C1'].alignment = CENTERALIGN
		ws['C1'].border = BOTTOMBORDER
		ws['D1'].value = "NH %"
		ws['D1'].font = HEADERFONT
		ws['D1'].alignment = CENTERALIGN
		ws['D1'].border = BOTTOMBORDER
		ws['E1'].value = "Refused"
		ws['E1'].font = HEADERFONT
		ws['E1'].alignment = CENTERALIGN
		ws['E1'].border = BOTTOMBORDER
		ws['F1'].value = "REF %"
		ws['F1'].font = HEADERFONT
		ws['F1'].alignment = CENTERALIGN
		ws['F1'].border = BOTTOMBORDER
		ws['G1'].value = "Moved"
		ws['G1'].font = HEADERFONT
		ws['G1'].alignment = CENTERALIGN
		ws['G1'].border = BOTTOMBORDER
		ws['H1'].value = "MVD %"
		ws['H1'].font = HEADERFONT
		ws['H1'].alignment = CENTERALIGN
		ws['H1'].border = BOTTOMBORDER
		ws['I1'].value = "Deceased"
		ws['I1'].font = HEADERFONT
		ws['I1'].alignment = CENTERALIGN
		ws['I1'].border = BOTTOMBORDER
		ws['J1'].value = "DEC %"
		ws['J1'].font = HEADERFONT
		ws['J1'].alignment = CENTERALIGN
		ws['J1'].border = BOTTOMBORDER
		ws['K1'].value = "Canvassed"
		ws['K1'].font = HEADERFONT
		ws['K1'].alignment = CENTERALIGN
		ws['K1'].border = BOTTOMBORDER
		ws['L1'].value = "CAN %"
		ws['L1'].font = HEADERFONT
		ws['L1'].alignment = CENTERALIGN
		ws['L1'].border = BOTTOMBORDER
		ws['M1'].value = "Call Back"
		ws['M1'].font = HEADERFONT
		ws['M1'].alignment = CENTERALIGN
		ws['M1'].border = BOTTOMBORDER
		ws['N1'].value = "CB %"
		ws['N1'].font = HEADERFONT
		ws['N1'].alignment = CENTERALIGN
		ws['N1'].border = BOTTOMBORDER
		ws['O1'].value = "Wrong Number"
		ws['O1'].font = HEADERFONT
		ws['O1'].alignment = CENTERALIGN
		ws['O1'].border = BOTTOMBORDER
		ws['P1'].value = "WB %"
		ws['P1'].font = HEADERFONT
		ws['P1'].alignment = CENTERALIGN
		ws['P1'].border = BOTTOMBORDER
		ws['Q1'].value = "Disconnected"
		ws['Q1'].font = HEADERFONT
		ws['Q1'].alignment = CENTERALIGN
		ws['Q1'].border = BOTTOMBORDER
		ws['R1'].value = "DC %"
		ws['R1'].font = HEADERFONT
		ws['R1'].alignment = CENTERALIGN
		ws['R1'].border = BOTTOMBORDER
		ws['S1'].value = "Busy"
		ws['S1'].font = HEADERFONT
		ws['S1'].alignment = CENTERALIGN
		ws['S1'].border = BOTTOMBORDER
		ws['T1'].value = "BSY %"
		ws['T1'].font = HEADERFONT
		ws['T1'].alignment = CENTERALIGN
		ws['T1'].border = BOTTOMBORDER
		ws['U1'].value = "Left Message"
		ws['U1'].font = HEADERFONT
		ws['U1'].alignment = CENTERALIGN
		ws['U1'].border = BOTTOMBORDER
		ws['V1'].value = "LM %"
		ws['V1'].font = HEADERFONT
		ws['V1'].alignment = CENTERALIGN
		ws['V1'].border = BOTTOMBORDER

		counter = 2

		for name in statsDict:
			rec = statsDict[name]
			c = str(counter)
			ws['A'+c].value = name
			ws['A'+c].alignment = RIGHTALIGN
			ws['A'+c].border = RIGHTBORDER
			ws['B'+c].value = rec[0]
			ws['B'+c].border = RIGHTBORDER
			ws['C'+c].value = rec[1]
			ws['D'+c].value = float(rec[2][0:-1])/100
			ws['D'+c].number_format = '0.00%'
			ws['D'+c].border = RIGHTBORDER
			ws['E'+c].value = rec[3]
			ws['F'+c].value = float(rec[4][0:-1])/100
			ws['F'+c].number_format = '0.00%'
			ws['F'+c].border = RIGHTBORDER
			ws['G'+c].value = rec[5]
			ws['H'+c].value = float(rec[4][0:-1])/100
			ws['H'+c].number_format = '0.00%'
			ws['H'+c].border = RIGHTBORDER
			ws['I'+c].value = rec[7]
			ws['J'+c].value = float(rec[8][0:-1])/100
			ws['J'+c].number_format = '0.00%'
			ws['J'+c].border = RIGHTBORDER
			ws['K'+c].value = rec[9]
			ws['L'+c].value = float(rec[10][0:-1])/100
			ws['L'+c].number_format = '0.00%'
			ws['L'+c].border = RIGHTBORDER
			ws['M'+c].value = rec[11]
			ws['N'+c].value = float(rec[12][0:-1])/100
			ws['N'+c].number_format = '0.00%'
			ws['N'+c].border = RIGHTBORDER
			ws['O'+c].value = rec[13]
			ws['P'+c].value = float(rec[14][0:-1])/100
			ws['P'+c].number_format = '0.00%'
			ws['P'+c].border = RIGHTBORDER
			ws['Q'+c].value = rec[15]
			ws['R'+c].value = float(rec[16][0:-1])/100
			ws['R'+c].number_format = '0.00%'
			ws['R'+c].border = RIGHTBORDER
			ws['S'+c].value = rec[17]
			ws['T'+c].value = float(rec[18][0:-1])/100
			ws['T'+c].number_format = '0.00%'
			ws['T'+c].border = RIGHTBORDER
			ws['U'+c].value = rec[19]
			ws['V'+c].value = float(rec[20][0:-1])/100
			ws['V'+c].number_format = '0.00%'
			ws['V'+c].border = RIGHTBORDER
			counter+=1

			if name == 'Averages':
				ws['A'+c].font = BOLDFONT
				ws['B'+c].border = TOPBOTTOMBORDER
				ws['B'+c].font = BOLDFONT
				ws['C'+c].border = TOPBOTTOMBORDER
				ws['C'+c].font = BOLDFONT
				ws['D'+c].border = TOPBOTTOMBORDER
				ws['D'+c].font = BOLDFONT
				ws['E'+c].border = TOPBOTTOMBORDER
				ws['E'+c].font = BOLDFONT
				ws['F'+c].border = TOPBOTTOMBORDER
				ws['F'+c].font = BOLDFONT
				ws['G'+c].border = TOPBOTTOMBORDER
				ws['G'+c].font = BOLDFONT
				ws['H'+c].border = TOPBOTTOMBORDER
				ws['H'+c].font = BOLDFONT
				ws['I'+c].border = TOPBOTTOMBORDER
				ws['I'+c].font = BOLDFONT
				ws['J'+c].border = TOPBOTTOMBORDER
				ws['J'+c].font = BOLDFONT
				ws['K'+c].border = TOPBOTTOMBORDER
				ws['K'+c].font = BOLDFONT
				ws['L'+c].border = TOPBOTTOMBORDER
				ws['L'+c].font = BOLDFONT
				ws['M'+c].border = TOPBOTTOMBORDER
				ws['M'+c].font = BOLDFONT
				ws['N'+c].border = TOPBOTTOMBORDER
				ws['N'+c].font = BOLDFONT
				ws['O'+c].border = TOPBOTTOMBORDER
				ws['O'+c].font = BOLDFONT
				ws['P'+c].border = TOPBOTTOMBORDER
				ws['P'+c].font = BOLDFONT
				ws['Q'+c].border = TOPBOTTOMBORDER
				ws['Q'+c].font = BOLDFONT
				ws['R'+c].border = TOPBOTTOMBORDER
				ws['R'+c].font = BOLDFONT
				ws['S'+c].border = TOPBOTTOMBORDER
				ws['S'+c].font = BOLDFONT
				ws['T'+c].border = TOPBOTTOMBORDER
				ws['T'+c].font = BOLDFONT
				ws['U'+c].font = BOLDFONT
				ws['U'+c].border = TOPBOTTOMBORDER
				ws['V'+c].font = BOLDFONT
				ws['V'+c].border = TOPBOTTOMBORDER

		unavailable = []
		for name in TEAMS[group]:
			if name in statsDict:
				pass
			else:
				unavailable.append(name)

		if len(unavailable) > 0:
			counter += 2

			ws['A'+str(counter)].value = "NO DATA FOR"
			ws['A'+str(counter)].alignment = RIGHTALIGN
			ws['A'+str(counter)].font = BOLDFONT

			for name in TEAMS[group]:
				if name in statsDict:
					pass
				else:
					counter += 1
					ws['A'+str(counter)].value = name

	def finishAverages(self,average,numWorked):
		### turns the row into its final form where it can be written into excel
		average = self.divideList(average,numWorked)
		newList = []
		for x,elem in enumerate(average):
			if x == 0:
				roundedNum = round(elem,2)
				newList.append(roundedNum)
			elif x%2 == 0:
				roundedNum = round(elem,2)
				newList.append(f'{roundedNum}%')
			else:
				roundedNum = round(elem,2)
				newList.append(roundedNum)
		return newList


	def divideList(self,listA,n):
		newList = []
		for elem in listA:
			newList.append(elem/n)
		return newList

	def addLists(self,listA,listB):
		result = []
		for x, item in enumerate(listA):
			if type(listB[x]) == str:
				result.append(float(listB[x][:-1]))
			else:
				result.append(item+listB[x])
		return result

	def changeName(self,team,wb,wsDicty):

		translationDict = {'Eight, Fvaw':'Katiuska Gutierrez',
						'Fifteen, Fvaw':'Jose Martinez',
						'Fourteen, Fvaw':'Yurima Saavedra',
						'Nine, Fvaw':'Ana Figueroa',
						'Six, Fvaw':'Yetsika Contreras',
						'Ten, Fvaw':'Francesca Pisani',
						'Thirteen, Fvaw':'Vicky Torrealba',
						'Twelve, Fvaw':'Yadhira Barrios',
						'Five, Fvaw':'Missouri Mancillia',
						'Four, Fvaw':'Angelica Diaz ',
						'Moreno Fuentes, Lazaraj':'Lazara Moreno',
						'Nineteen, Fvaw':'Lizyasiri Perez',
						'One, Fvaw':'Maria F',
						'Sisepuede2020, Fvawcanvasser21@gmai':'Maria Gonzalez',
						'Twenty, Fvaw':'Rudy Camacho',
						'Twentytwo, Fvaw':'Carmina Redonet',
						'Two, Fvaw':'Carolina Perez'}


		ws = wb.worksheets[0]


		names = wb.sheetnames

		for x,sheet in enumerate(wb.worksheets):
			if sheet.title == "Team Data":
				pass
			else:
				sheet.title = translationDict[sheet.title]

		loopNum = len(names)-1

		ws.insert_cols(1)
		ws.column_dimensions['B'].width = 15
		starting = 2

		ws['A1'].value = "Canvasser"
		ws['B1'].value = "Username"
		ws['A1'].font = HEADERFONT
		ws['B1'].font = HEADERFONT
		ws['A1'].alignment = RIGHTALIGN
		ws['B1'].alignment = LEFTALIGN

		for x in range(loopNum):
			y = f'A{starting}'
			nameCoords = f'B{starting}'
			name = ws[nameCoords].value
			ws[nameCoords].alignment = LEFTALIGN
			print(name)
			if name in translationDict:
				ws[y].value = translationDict[name]
				ws[y].alignment = RIGHTALIGN
				ws[y].font = BOLDFONT		
			starting += 1

		ws.column_dimensions['A'].width = 22
		ws.column_dimensions['B'].width = 22
		ws.column_dimensions['C'].width = 8
		ws.column_dimensions['D'].width = 11
		ws.column_dimensions['E'].width = 8
		ws.column_dimensions['F'].width = 10
		ws.column_dimensions['G'].width = 8
		ws.column_dimensions['H'].width = 10
		ws.column_dimensions['I'].width = 8
		ws.column_dimensions['J'].width = 10
		ws.column_dimensions['K'].width = 8
		ws.column_dimensions['L'].width = 12
		ws.column_dimensions['M'].width = 8
		ws.column_dimensions['N'].width = 12
		ws.column_dimensions['O'].width = 8
		ws.column_dimensions['P'].width = 16
		ws.column_dimensions['Q'].width = 8
		ws.column_dimensions['R'].width = 14
		ws.column_dimensions['S'].width = 8
		ws.column_dimensions['T'].width = 8
		ws.column_dimensions['U'].width = 8
		ws.column_dimensions['V'].width = 14
		ws.column_dimensions['W'].width = 8

	def populateDetail(self,ws,rows):
		### takes in worksheet and data
		### and populates that worksheet
		### with that data
		ws.column_dimensions['A'].hidden = True
		ws.sheet_view.showGridLines= False

		ws.column_dimensions['B'].width = 24
		ws.column_dimensions['C'].width = 11.5
		ws.column_dimensions['D'].width = 5
		ws.column_dimensions['E'].width = 5
		ws.column_dimensions['F'].width = 5
		ws.column_dimensions['G'].width = 5
		ws.column_dimensions['H'].width = 5
		ws.column_dimensions['I'].width = 5

		ws['A1'].value = "Address"
		ws['A1'].font = HEADERFONT
		ws['A1'].border = BOTTOMBORDER
		ws['B1'].value = "Person"
		ws['B1'].font = HEADERFONT
		ws['B1'].border = BOTTOMBORDER
		ws['C1'].value = "Time"
		ws['C1'].font = HEADERFONT
		ws['C1'].border = BOTTOMBORDER
		ws['D1'].value = "NH"
		ws['D1'].font = HEADERFONT
		ws['D1'].border = BOTTOMBORDER
		ws['E1'].value = "Ref"
		ws['E1'].font = HEADERFONT
		ws['E1'].border = BOTTOMBORDER
		ws['F1'].value = "Mvd"
		ws['F1'].font = HEADERFONT
		ws['F1'].border = BOTTOMBORDER
		ws['G1'].value = "Dec"
		ws['G1'].font = HEADERFONT
		ws['G1'].border = BOTTOMBORDER
		ws['H1'].value = "Canv"
		ws['H1'].font = HEADERFONT
		ws['H1'].border = BOTTOMBORDER
		ws['I1'].border = BOTTOMBORDER
		ws['J1'].border = BOTTOMBORDER
		ws['K1'].border = BOTTOMBORDER

		ws['N2'].value = "WN = Wrong Number"
		ws['N3'].value = "CB = Call Back"
		ws['N4'].value = "DC = Disconnected"
		ws['N5'].value = "BS = Busy"
		ws['N6'].value = "LM = Left Message"

		c = 2 ## row location index thing
		for row in rows:
			ws['A'+str(c)].value = row[0]
			ws['A'+str(c)].border = ALLBORDER
			ws['B'+str(c)].value = row[1]
			ws['B'+str(c)].border = ALLBORDER
			# print(row[2])
			x = row[2].strftime('%I:%M:00 %p')
			# print(x)
			ws['C'+str(c)].value = x
			ws['C'+str(c)].border = ALLBORDER
			ws['D'+str(c)].value = row[3]
			ws['D'+str(c)].font = BOLDFONT
			ws['D'+str(c)].border = ALLBORDER
			ws['D'+str(c)].alignment = CENTERALIGN
			ws['E'+str(c)].value = row[4]
			ws['E'+str(c)].border = ALLBORDER
			ws['E'+str(c)].font = BOLDFONT
			ws['E'+str(c)].alignment = CENTERALIGN
			ws['F'+str(c)].value = row[5]
			ws['F'+str(c)].font = BOLDFONT
			ws['F'+str(c)].alignment = CENTERALIGN
			ws['F'+str(c)].border = ALLBORDER
			ws['G'+str(c)].value = row[6]
			ws['G'+str(c)].font = BOLDFONT
			ws['G'+str(c)].alignment = CENTERALIGN
			ws['G'+str(c)].border = ALLBORDER
			ws['H'+str(c)].value = row[7]
			ws['H'+str(c)].font = BOLDFONT
			ws['H'+str(c)].alignment = CENTERALIGN
			ws['H'+str(c)].border = ALLBORDER
			ws['J'+str(c)].border = ALLBORDER
			ws['K'+str(c)].border = ALLBORDER
			if row[8] != 0:
				if row[8] == "Call Back":
					ws['I'+str(c)].value = "CB"
				elif row[8] == "Busy":
					ws['I'+str(c)].value = "BS"
				elif row[8] == "Wrong Number":
					ws['I'+str(c)].value = "WN"
				elif row[8] == "Disconnected":
					ws['I'+str(c)].value = "DC"
				elif row[8] == "Left Message":
					ws['I'+str(c)].value = "LM"
				else:
					pass
					# print(f'row[8] {row[8]}')
			ws['I'+str(c)].font = BOLDFONT
			ws['I'+str(c)].alignment = CENTERALIGN
			ws['I'+str(c)].border = ALLBORDER
			if c%2 == 0:
				ws['A'+str(c)].fill = DARKGRAYFILL
				ws['B'+str(c)].fill = DARKGRAYFILL
				ws['C'+str(c)].fill = DARKGRAYFILL
				ws['D'+str(c)].fill = DARKGRAYFILL
				ws['E'+str(c)].fill = DARKGRAYFILL
				ws['F'+str(c)].fill = DARKGRAYFILL
				ws['G'+str(c)].fill = DARKGRAYFILL
				ws['H'+str(c)].fill = DARKGRAYFILL
				ws['I'+str(c)].fill = DARKGRAYFILL
				ws['J'+str(c)].fill = DARKGRAYFILL
				ws['K'+str(c)].fill = DARKGRAYFILL


			c += 1

	def getIndividualStats(self,rows):
		total = len(rows)

		nH = 0
		ref = 0
		mvd = 0
		decl = 0
		canv = 0
		cB = 0
		wN = 0
		dC = 0
		bsy = 0
		lM = 0

		for row in rows:
			if row[3] == 'X':
				nH += 1
			if row[4] == "X":
				ref += 1
			if row[5] == "X":
				mvd += 1
			if row[6] == "X":
				decl += 1
			if row[7] == "X":
				canv += 1
			if row[8] == "Wrong Number":
				wN += 1
			if row[8] == "Call Back":
				cB += 1
			if row[8] == "Disconnected":
				dC += 1
			if row[8] == "Left Message":
				lM += 1

		nHD = str(round((nH/total*100),2))+"%"
		refD = str(round((ref/total*100),2))+"%"
		mvdD = str(round((mvd/total*100),2))+"%"
		declD = str(round((decl/total*100),2))+"%"
		canvD = str(round((canv/total*100),2))+"%"
		cBD = str(round((cB/total*100),2))+"%"
		wND = str(round((wN/total*100),2))+"%"
		dCD = str(round((dC/total*100),2))+"%"
		bsyD = str(round((bsy/total*100),2))+"%"
		lMD = str(round((lM/total*100),2))+"%"
		stats = [
			total,nH,nHD,ref,refD,mvd,mvdD,decl,declD,
			canv,canvD,cB,cBD,wN,wND,dC,dCD,bsy,bsyD,
			lM,lMD
				]
		return stats




	def saveFile(self,group,wb):
		date = datetime.today().strftime('%m-%d-%y')
		wb.save(f'{group} {date}.xlsx')



		


excelCreator = ExcelCreator('finalDict2.p')
for team in TEAMS:
	excelCreator.createFile(team)

